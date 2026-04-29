import io
import re
import cv2
import numpy as np
from typing import List, Tuple

from paddleocr import PaddleOCR
from pdf2image import convert_from_bytes
from openpyxl import load_workbook
from docx import Document
import easyocr


# ============================================================
# 0. PaddleOCR モデル生成（A/C）
# ============================================================

def create_paddle_model(model_type: str):
    """
    model_type:
        "A" → 最速（PP-OCRv5_mobile_rec）
        "B" → Vision（ここでは None を返す）
        "C" → fallback（Multilingual）
    """

    if model_type == "A":
        # ★ あなたの環境に合わせてパスを設定
        return PaddleOCR(
            lang="japan",
            use_angle_cls=True,
            rec_model_dir="models/japan/v5_mobile"
        )

    elif model_type == "C":
        return PaddleOCR(
            lang="japan",
            use_angle_cls=True
        )

    elif model_type == "B":
        # Vision OCR は app.py 側で扱うため None を返す
        return None

    raise ValueError(f"Unknown model_type: {model_type}")


# ============================================================
# 1. 前処理
# ============================================================

def preprocess_image(
    img: np.ndarray,
    adv_binarize: bool,
    adv_deskew: bool,
    adv_denoise: bool,
    adv_contrast: bool,
    is_handwritten: bool
) -> np.ndarray:

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    if adv_binarize:
        _, th = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)
    else:
        _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    if adv_deskew:
        coords = np.column_stack(np.where(th < 255))
        if coords.size > 0:
            angle = cv2.minAreaRect(coords)[-1]
            if angle < -45:
                angle = -(90 + angle)
            else:
                angle = -angle
            (h, w) = th.shape
            M = cv2.getRotationMatrix2D((w // 2, h // 2), angle, 1.0)
            th = cv2.warpAffine(th, M, (w, h), flags=cv2.INTER_CUBIC)

    if adv_denoise:
        th = cv2.GaussianBlur(th, (3, 3), 0)

    if adv_contrast:
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        th = clahe.apply(th)

    if not is_handwritten:
        th = cv2.resize(th, None, fx=0.85, fy=0.85, interpolation=cv2.INTER_LINEAR)
    else:
        th = cv2.resize(th, None, fx=1.5, fy=1.5, interpolation=cv2.INTER_LINEAR)

    return th


# ============================================================
# 2. 行分割
# ============================================================

def segment_lines(img: np.ndarray, line_threshold: float) -> List[np.ndarray]:
    projection = np.sum(img < 128, axis=1)
    if np.max(projection) == 0:
        return []

    threshold = np.max(projection) * line_threshold

    lines = []
    start = None

    for i, val in enumerate(projection):
        if val > threshold and start is None:
            start = i
        elif val <= threshold and start is not None:
            end = i
            line_img = img[start:end, :]
            if line_img.shape[0] > 8:
                lines.append(line_img)
            start = None

    if start is not None:
        line_img = img[start:, :]
        if line_img.shape[0] > 8:
            lines.append(line_img)

    return lines


# ============================================================
# 3. 後処理（手書き時のみ）
# ============================================================

def postprocess_text(text: str) -> str:
    text = text.replace("O", "0").replace("o", "0")
    text = text.replace("l", "1").replace("I", "1")
    text = text.replace("S", "5")
    text = re.sub(r"(\d+)[,\.](\d{3})", r"\1\2", text)
    return text


# ============================================================
# 4. 言語判定
# ============================================================

def detect_language(text: str) -> str:
    jp_chars = re.findall(r"[ぁ-んァ-ン一-龯ー]", text)
    if len(jp_chars) >= 5 and len(jp_chars) / max(len(text), 1) > 0.10:
        return "JP"
    return "EN"


# ============================================================
# 5. PDF → 画像
# ============================================================

def pdf_to_images(file_bytes: bytes) -> List[np.ndarray]:
    pil_pages = convert_from_bytes(file_bytes, dpi=150)
    cv_pages = []
    for page in pil_pages:
        page = page.convert("RGB")
        np_img = np.array(page)
        cv_img = cv2.cvtColor(np_img, cv2.COLOR_RGB2BGR)
        cv_pages.append(cv_img)
    return cv_pages


# ============================================================
# 6. OCR（手書き判定 + モデル切り替え）
# ============================================================

ocr_jp_hand = easyocr.Reader(['ja', 'en'], gpu=False)
ocr_en = easyocr.Reader(['en'], gpu=False)


def ocr_page(
    img: np.ndarray,
    lang: str,
    paddle_model,
    adv_binarize: bool,
    adv_deskew: bool,
    adv_denoise: bool,
    adv_contrast: bool,
    line_threshold: float
) -> Tuple[str, bool]:

    processed = preprocess_image(img, adv_binarize, adv_deskew, adv_denoise, adv_contrast, False)
    lines = segment_lines(processed, line_threshold)

    texts = []
    confidences = []

    for line in lines:
        if lang == "EN":
            result = ocr_en.readtext(line, detail=1)
            for _, txt, conf in result:
                confidences.append(conf)
                texts.append(txt)
        else:
            result = paddle_model.ocr(line, cls=True)
            for res in result:
                txt = res[1][0]
                conf = res[1][1]
                confidences.append(conf)
                texts.append(txt)

    is_handwritten = False
    if confidences:
        low_conf = [c for c in confidences if c < 0.70]
        if len(low_conf) / len(confidences) > 0.40:
            is_handwritten = True

    if lang == "JP" and is_handwritten:
        processed = preprocess_image(img, adv_binarize, adv_deskew, adv_denoise, adv_contrast, True)
        lines = segment_lines(processed, line_threshold)
        texts = []
        for line in lines:
            result = ocr_jp_hand.readtext(line, detail=1)
            for _, txt, _ in result:
                texts.append(postprocess_text(txt))

    return "\n".join(texts), is_handwritten


# ============================================================
# 7. Excel
# ============================================================

def extract_excel(file_bytes: bytes) -> str:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    texts = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join([str(c) for c in row if c is not None])
            if row_text.strip():
                texts.append(row_text)
    return "\n".join(texts)


# ============================================================
# 8. Word
# ============================================================

def extract_docx(file_bytes: bytes) -> str:
    doc = Document(io.BytesIO(file_bytes))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


# ============================================================
# 9. メイン（lang を返すように修正済み）
# ============================================================

def extract_raw_content(
    file,
    model_type="A",
    adv_binarize=False,
    adv_deskew=False,
    adv_denoise=True,
    adv_contrast=True,
    line_threshold=0.08
):
    raw_bytes = file.read()
    name = getattr(file, "name", "").lower()

    sample = raw_bytes[:3000].decode("utf-8", errors="ignore")
    lang = detect_language(sample)

    paddle_model = create_paddle_model(model_type)

    # PDF
    if raw_bytes.startswith(b"%PDF") or name.endswith(".pdf"):
        pages = pdf_to_images(raw_bytes)
        all_texts = []
        any_handwritten = False

        for page in pages:
            page_text, is_hand = ocr_page(
                page, lang, paddle_model,
                adv_binarize, adv_deskew, adv_denoise, adv_contrast, line_threshold
            )
            if page_text.strip():
                all_texts.append(page_text)
            if is_hand:
                any_handwritten = True

        return "\n\n".join(all_texts), "pdf", any_handwritten, lang

    # Excel
    if name.endswith(".xlsx"):
        return extract_excel(raw_bytes), "excel", False, lang

    # Word
    if name.endswith(".docx"):
        return extract_docx(raw_bytes), "docx", False, lang

    # Text
    if name.endswith(".txt"):
        try:
            text = raw_bytes.decode("utf-8", errors="ignore")
        except:
            text = raw_bytes.decode("cp932", errors="ignore")
        return text, "text", False, lang

    # Image
    np_img = np.frombuffer(raw_bytes, np.uint8)
    img = cv2.imdecode(np_img, cv2.IMREAD_COLOR)
    if img is None:
        try:
            text = raw_bytes.decode("utf-8", errors="ignore")
        except:
            text = ""
        return text, "unknown", False, lang

    page_text, is_hand = ocr_page(
        img, lang, paddle_model,
        adv_binarize, adv_deskew, adv_denoise, adv_contrast, line_threshold
    )
    return page_text, "image", is_hand, lang
